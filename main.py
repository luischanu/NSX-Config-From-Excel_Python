import openpyxl
import NSX_GoGet
import NSX_v

"""
Thoughts:
    1) Could create a separate Class for each type of NSX object, then instantiate as needed.  At that point, add each object to the list (in place of single_object.  Benefits include:
        a) Permitting easy conversion of text to INT's for respective fields
        b) Bounds checking for each field, when required.  For example, PrefixLength needs to be between 0 and 32 for IPv4.
    2) Currently, all values are strings.   For some items, like PrefixLength, those should be integers.
"""


def print_header():
    print("----------------------------------------------")
    print("    Read Excel Spreadsheet Into Dictionary    ")
    print("----------------------------------------------")


def excel_to_dict(XLSFile):
    workbook_data = {}

    workbook = openpyxl.load_workbook(filename=XLSFile, read_only=True)

    # Iterate through each worksheet in the workbook
    for worksheet in workbook:
        # Skip processing the worksheet if name starts with an underscore
        if worksheet.title[0] == "_":
            continue

        # The worksheet name is used as the key for the workbook dict, so we need to get it for later use.
        worksheet_name = worksheet.title
        print("Worksheet: {}".format(worksheet_name))
        sheet_data = []

        # Iterate through each of the rows in the spreadsheet
        for row in worksheet.iter_rows(row_offset=1):
            single_object = {}

            # Get the data from each cell in the row.  Use index to handle references to empty cells.
            for col_idx, cell in enumerate(row, 1):
                # Obtain the header of this cell, as it's used as the key in the single_object Dict
                # Using COL_IDX instead of referencing cell.column because empty cells have no properties, and thus, error out.
                header = worksheet.cell(row=1, column=col_idx).value

                # Check for empty cell.  If cell has contents, then store it, else store empty string.
                if cell.value:
                    single_object[header] = str(cell.value)
                else:
                    single_object[header] = ""

            # Add row of data from spreadsheet to sheet_data list
            print("      single_object: {}".format(single_object))
            sheet_data.append(single_object)

        # Add this sheet to the Dict, with the worksheet name being the key
        workbook_data[worksheet_name] = sheet_data

    # All done with XLSFile, so let's close it.
    workbook.close()

    # Return Dict of workbook data with the workbook names as the key values
    return workbook_data


def main():
    nsxmgr = NSX_v.NSXManager("SFO-vCenter.Lab.Local", "Administrator@vSphere.Local","Hello123!")
    XLSFilename = "VirtualEnvironment.xlsx"

    print_header()

    XLSData = excel_to_dict(XLSFilename)
    print(XLSData)

    print("=====================================================================================================================================================================")
    print(XLSData["Infrastructure"])
    print("=====================================================================================================================================================================")
    print(XLSData["Infrastructure"][0])
    print("=====================================================================================================================================================================")
    print((XLSData["Infrastructure"][0])["vCenterServer"])

    print("=====================================================================================================================================================================")


    #NSX Manager
    nsxmgr = NSX_v.NSXManager(
        (XLSData["Infrastructure"][0])["vCenterServer"],        # vCenterServer Address
        (XLSData["Infrastructure"][0])["vCenter_Admin"],        # vCenter Admin Username
        (XLSData["Infrastructure"][0])["vCenter_Password"]      # vCenter Admin Password
    )

    print("NSX Manager: {}".format(nsxmgr))
    print("=====================================================================================================================================================================")




    tz_list = XLSData["TransportZones"]
    print(f"NSX Manager: {nsxmgr}")

    print("TransportZone List: {}, ObjectID: ".format(tz_list))

    for tz_dict in tz_list:
        print("TZ Name: {},  TZ ObjectID: {}".format(tz_dict['Name'], NSX_GoGet.transportzone_oid(nsxmgr, tz_dict['Name'])))





if __name__ == "__main__":
    main()
